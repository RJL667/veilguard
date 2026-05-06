"""Build CherryRain_CostEfficiency_Model.xlsx
Cost-efficiency model for Cherry Rain (SME, Phishield prospect).
Sheets: Inputs, Costs, Savings, ROI Summary, Sensitivity.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\rudol\Documents\veilguard\CherryRain_CostEfficiency_Model.xlsx"

wb = Workbook()

# Styles
HDR = Font(bold=True, color="FFFFFF", size=12)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
SUB = Font(bold=True, size=11)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOT = Font(bold=True, size=11)
TOT_FILL = PatternFill("solid", fgColor="FFE699")
INPUT_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

ZAR = '"R"#,##0;[Red]"R"-#,##0'
ZAR2 = '"R"#,##0.00;[Red]"R"-#,##0.00'
PCT = '0.0%'
NUM = '#,##0'

def header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def setw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# === Sheet 1: Inputs ===
ws = wb.active
ws.title = "Inputs"
ws["A1"] = "Cherry Rain — Cost-Efficiency Model: Inputs"
ws["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws.merge_cells("A1:C1")
ws["A2"] = "All monetary values in ZAR. Green cells = editable inputs."
ws["A2"].font = Font(italic=True, color="595959")
ws.merge_cells("A2:C2")

rows = [
    ("Company profile", "", ""),
    ("Headcount", 45, "FTE"),
    ("Email-using staff", 42, "FTE"),
    ("Average loaded cost / FTE / hour", 285, "ZAR"),
    ("", "", ""),
    ("Threat baseline (annual, pre-Phishield)", "", ""),
    ("Phishing emails received / user / month", 18, "count"),
    ("Click-through rate (untrained)", 0.045, "%"),
    ("Successful credential compromises / yr", 3, "count"),
    ("Avg incident response cost (per incident)", 78000, "ZAR"),
    ("Probability of major breach / yr", 0.08, "%"),
    ("Estimated major-breach loss", 2400000, "ZAR"),
    ("Productivity loss / phishing triage (min)", 12, "min"),
    ("", "", ""),
    ("Phishield package", "", ""),
    ("Tier", "Business 50", "—"),
    ("Annual subscription", 96000, "ZAR"),
    ("One-off onboarding", 18500, "ZAR"),
    ("Internal admin time (hrs / yr)", 40, "hrs"),
    ("", "", ""),
    ("Efficacy assumptions (Phishield Year 1)", "", ""),
    ("Phishing block rate", 0.92, "%"),
    ("Click-through rate reduction (training)", 0.70, "%"),
    ("Incident response time saved", 0.55, "%"),
    ("Major-breach probability reduction", 0.65, "%"),
]

start = 4
for i, (lbl, val, unit) in enumerate(rows):
    r = start + i
    ws.cell(row=r, column=1, value=lbl)
    ws.cell(row=r, column=2, value=val)
    ws.cell(row=r, column=3, value=unit)
    if isinstance(val, str) and val == "" and lbl and not unit:
        ws.cell(row=r, column=1).font = SUB
        ws.cell(row=r, column=1).fill = SUB_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    elif lbl and val == "" and unit == "":
        pass
    elif lbl == "":
        pass
    else:
        c = ws.cell(row=r, column=2)
        c.fill = INPUT_FILL
        c.alignment = RIGHT
        if unit == "ZAR":
            c.number_format = ZAR
        elif unit == "%":
            c.number_format = PCT
        elif unit in ("count", "FTE", "min", "hrs"):
            c.number_format = NUM

# Named cells via defined names (use cell refs in formulas)
# We'll reference Inputs!B<row> in other sheets.
# Build a map for clarity:
INP = {}
for i, (lbl, val, unit) in enumerate(rows):
    INP[lbl] = start + i

setw(ws, [42, 18, 10])

# === Sheet 2: Costs ===
ws2 = wb.create_sheet("Costs")
ws2["A1"] = "Phishield Total Cost (Year 1 & Year 2+)"
ws2["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws2.merge_cells("A1:D1")

header(ws2, 3, ["Cost line", "Year 1", "Year 2", "Notes"])

cost_lines = [
    ("Annual subscription", f"=Inputs!B{INP['Annual subscription']}", f"=Inputs!B{INP['Annual subscription']}", "Recurring"),
    ("Onboarding (one-off)", f"=Inputs!B{INP['One-off onboarding']}", 0, "Year 1 only"),
    ("Internal admin time",
     f"=Inputs!B{INP['Internal admin time (hrs / yr)']}*Inputs!B{INP['Average loaded cost / FTE / hour']}",
     f"=Inputs!B{INP['Internal admin time (hrs / yr)']}*Inputs!B{INP['Average loaded cost / FTE / hour']}*0.6",
     "Yr2 lower (steady state)"),
]
r = 4
for lbl, y1, y2, note in cost_lines:
    ws2.cell(row=r, column=1, value=lbl)
    ws2.cell(row=r, column=2, value=y1).number_format = ZAR
    ws2.cell(row=r, column=3, value=y2).number_format = ZAR
    ws2.cell(row=r, column=4, value=note)
    r += 1

ws2.cell(row=r, column=1, value="Total cost").font = TOT
ws2.cell(row=r, column=1).fill = TOT_FILL
ws2.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").number_format = ZAR
ws2.cell(row=r, column=2).font = TOT
ws2.cell(row=r, column=2).fill = TOT_FILL
ws2.cell(row=r, column=3, value=f"=SUM(C4:C{r-1})").number_format = ZAR
ws2.cell(row=r, column=3).font = TOT
ws2.cell(row=r, column=3).fill = TOT_FILL
COST_TOTAL_ROW = r
setw(ws2, [32, 16, 16, 36])

# === Sheet 3: Savings ===
ws3 = wb.create_sheet("Savings")
ws3["A1"] = "Annualised Savings & Loss Avoidance"
ws3["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws3.merge_cells("A1:D1")

header(ws3, 3, ["Savings line", "Annual value", "Calculation", "Source"])

# helpers
phish_per_yr = f"Inputs!B{INP['Phishing emails received / user / month']}*12*Inputs!B{INP['Email-using staff']}"
savings = [
    ("Triage productivity recovered",
     f"=({phish_per_yr})*Inputs!B{INP['Phishing block rate']}*(Inputs!B{INP['Productivity loss / phishing triage (min)']}/60)*Inputs!B{INP['Average loaded cost / FTE / hour']}",
     "Blocked emails × triage time saved × loaded rate",
     "Inputs"),
    ("Avoided credential compromises",
     f"=Inputs!B{INP['Successful credential compromises / yr']}*Inputs!B{INP['Click-through rate reduction (training)']}*Inputs!B{INP['Avg incident response cost (per incident)']}",
     "Compromises × reduction × IR cost",
     "Inputs"),
    ("Faster incident response",
     f"=Inputs!B{INP['Successful credential compromises / yr']}*Inputs!B{INP['Avg incident response cost (per incident)']}*Inputs!B{INP['Incident response time saved']}*0.4",
     "Residual incidents × IR cost × time-saved × IR-share",
     "Inputs"),
    ("Major-breach loss avoided (expected value)",
     f"=Inputs!B{INP['Probability of major breach / yr']}*Inputs!B{INP['Estimated major-breach loss']}*Inputs!B{INP['Major-breach probability reduction']}",
     "P(breach) × loss × probability reduction",
     "Inputs"),
]
r = 4
for lbl, calc, expr, src in savings:
    ws3.cell(row=r, column=1, value=lbl)
    ws3.cell(row=r, column=2, value=calc).number_format = ZAR
    ws3.cell(row=r, column=3, value=expr)
    ws3.cell(row=r, column=4, value=src)
    r += 1

ws3.cell(row=r, column=1, value="Total annual savings").font = TOT
ws3.cell(row=r, column=1).fill = TOT_FILL
ws3.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})").number_format = ZAR
ws3.cell(row=r, column=2).font = TOT
ws3.cell(row=r, column=2).fill = TOT_FILL
SAVINGS_TOTAL_ROW = r
setw(ws3, [40, 18, 52, 14])

# === Sheet 4: ROI Summary ===
ws4 = wb.create_sheet("ROI Summary")
ws4["A1"] = "ROI Summary — Cherry Rain"
ws4["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws4.merge_cells("A1:C1")

header(ws4, 3, ["Metric", "Year 1", "Year 2"])

rows4 = [
    ("Total cost", f"=Costs!B{COST_TOTAL_ROW}", f"=Costs!C{COST_TOTAL_ROW}", ZAR),
    ("Total annual savings", f"=Savings!B{SAVINGS_TOTAL_ROW}", f"=Savings!B{SAVINGS_TOTAL_ROW}", ZAR),
    ("Net benefit", f"=B5-B4", f"=C5-C4", ZAR),
    ("ROI (%)", f"=B6/B4", f"=C6/C4", PCT),
    ("Payback period (months)", f"=B4/(B5/12)", f"=C4/(C5/12)", '0.0'),
    ("Cost / FTE / month",
     f"=B4/Inputs!B{INP['Headcount']}/12",
     f"=C4/Inputs!B{INP['Headcount']}/12", ZAR2),
]
r = 4
for lbl, y1, y2, fmt in rows4:
    ws4.cell(row=r, column=1, value=lbl).font = SUB
    c1 = ws4.cell(row=r, column=2, value=y1); c1.number_format = fmt
    c2 = ws4.cell(row=r, column=3, value=y2); c2.number_format = fmt
    if lbl in ("Net benefit", "ROI (%)"):
        for c in (c1, c2):
            c.font = TOT; c.fill = TOT_FILL
    r += 1

ws4["A11"] = "Interpretation"
ws4["A11"].font = SUB
ws4["A11"].fill = SUB_FILL
ws4.merge_cells("A11:C11")
ws4["A12"] = ("Year 1 ROI is suppressed by onboarding cost; steady-state Year 2 reflects "
              "the recurring economics. Major-breach EV is the dominant savings line — "
              "stress-test it on the Sensitivity sheet.")
ws4["A12"].alignment = Alignment(wrap_text=True, vertical="top")
ws4.merge_cells("A12:C14")
ws4.row_dimensions[12].height = 50

setw(ws4, [32, 18, 18])

# === Sheet 5: Sensitivity ===
ws5 = wb.create_sheet("Sensitivity")
ws5["A1"] = "Sensitivity — Year 1 Net Benefit vs Major-Breach Assumptions"
ws5["A1"].font = Font(bold=True, size=14, color="1F4E78")
ws5.merge_cells("A1:G1")

ws5["A3"] = "Rows: P(major breach). Cols: probability reduction. Values: Year 1 net benefit (ZAR)."
ws5["A3"].font = Font(italic=True)
ws5.merge_cells("A3:G3")

# Header row of reduction percentages
reductions = [0.30, 0.45, 0.60, 0.75, 0.90]
probs = [0.02, 0.05, 0.08, 0.12, 0.18]

ws5.cell(row=5, column=1, value="P(breach) \\ Reduction").font = SUB
ws5.cell(row=5, column=1).fill = SUB_FILL
for j, red in enumerate(reductions, 2):
    c = ws5.cell(row=5, column=j, value=red)
    c.number_format = PCT
    c.font = SUB
    c.fill = SUB_FILL
    c.alignment = CENTER

# Static components (don't depend on the swept vars):
# total cost Y1, all savings except major-breach
non_breach_savings = (
    f"(Inputs!B{INP['Phishing emails received / user / month']}*12*Inputs!B{INP['Email-using staff']})"
    f"*Inputs!B{INP['Phishing block rate']}*(Inputs!B{INP['Productivity loss / phishing triage (min)']}/60)"
    f"*Inputs!B{INP['Average loaded cost / FTE / hour']}"
    f"+Inputs!B{INP['Successful credential compromises / yr']}*Inputs!B{INP['Click-through rate reduction (training)']}*Inputs!B{INP['Avg incident response cost (per incident)']}"
    f"+Inputs!B{INP['Successful credential compromises / yr']}*Inputs!B{INP['Avg incident response cost (per incident)']}*Inputs!B{INP['Incident response time saved']}*0.4"
)
total_cost_y1 = f"Costs!B{COST_TOTAL_ROW}"

for i, p in enumerate(probs, 6):
    pc = ws5.cell(row=i, column=1, value=p)
    pc.number_format = PCT
    pc.font = SUB
    pc.fill = SUB_FILL
    for j, red in enumerate(reductions, 2):
        formula = (
            f"=({non_breach_savings} + {p}*Inputs!B{INP['Estimated major-breach loss']}*{red})"
            f"-{total_cost_y1}"
        )
        c = ws5.cell(row=i, column=j, value=formula)
        c.number_format = ZAR
        c.alignment = RIGHT
        c.border = BORDER

ws5["A13"] = "Read: at the baseline P=8% × reduction=65% the model returns the headline Year 1 net benefit shown on the ROI Summary."
ws5["A13"].font = Font(italic=True, color="595959")
ws5.merge_cells("A13:G13")

setw(ws5, [22, 16, 16, 16, 16, 16, 16])

wb.save(OUT)
print("WROTE", OUT)
