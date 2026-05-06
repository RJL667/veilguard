import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Model"

P = 50_000_000
T = 5
LOAN_RATE = 0.115
BOND_COUPON = 0.0925
BOND_ISS = 750_000
BOND_ADMIN = 60_000
LOAN_FEE_RATE = 0.015
LOAN_ADMIN = 25_000
DR = 0.10

loan_fee = P * LOAN_FEE_RATE
loan_int = P * LOAN_RATE
loan_annual = loan_int + LOAN_ADMIN
loan_total = loan_fee + loan_int * T + LOAN_ADMIN * T
# NPV: fee at t0, annual cost t1..t5
loan_npv = loan_fee + sum(loan_annual / (1 + DR) ** t for t in range(1, T + 1))

bond_coupon = P * BOND_COUPON
bond_annual = bond_coupon + BOND_ADMIN
bond_total = BOND_ISS + bond_coupon * T + BOND_ADMIN * T
bond_npv = BOND_ISS + sum(bond_annual / (1 + DR) ** t for t in range(1, T + 1))

nom_save = loan_total - bond_total
npv_save = loan_npv - bond_npv
save_pct = npv_save / loan_npv

rows = [
    ["CHERRY RAIN PROPERTIES — COST EFFICIENCY MODEL", "", ""],
    ["Structured Bond vs Direct Loan", "", ""],
    ["", "", ""],
    ["INPUTS", "Value", "Notes"],
    ["Principal (ZAR)", P, "Capital required"],
    ["Tenor (years)", T, "Same tenor both options"],
    ["Direct Loan Rate", LOAN_RATE, "Fixed annual"],
    ["Bond Coupon", BOND_COUPON, "Fixed annual"],
    ["Bond Issuance Costs", BOND_ISS, "Once-off"],
    ["Bond Annual Admin", BOND_ADMIN, "Trustee/listing"],
    ["Loan Origination Fee Rate", LOAN_FEE_RATE, "% of principal"],
    ["Loan Annual Admin", LOAN_ADMIN, "Bank fees"],
    ["Discount Rate", DR, "For NPV"],
    ["", "", ""],
    ["DIRECT LOAN", "Amount (ZAR)", ""],
    ["Origination Fee", loan_fee, "=Principal*1.5%"],
    ["Annual Interest", loan_int, "=Principal*11.5%"],
    ["Annual Admin", LOAN_ADMIN, ""],
    ["Annual Cost", loan_annual, "Interest+Admin"],
    ["Total Interest (5y)", loan_int * T, ""],
    ["Total Admin (5y)", LOAN_ADMIN * T, ""],
    ["Total Nominal Cost", loan_total, "Origination+Interest+Admin"],
    ["NPV of Costs", round(loan_npv, 2), "Discounted at 10%"],
    ["", "", ""],
    ["STRUCTURED BOND", "Amount (ZAR)", ""],
    ["Issuance Costs", BOND_ISS, "Once-off"],
    ["Annual Coupon", bond_coupon, "=Principal*9.25%"],
    ["Annual Admin", BOND_ADMIN, ""],
    ["Annual Cost", bond_annual, "Coupon+Admin"],
    ["Total Coupon (5y)", bond_coupon * T, ""],
    ["Total Admin (5y)", BOND_ADMIN * T, ""],
    ["Total Nominal Cost", bond_total, "Issuance+Coupon+Admin"],
    ["NPV of Costs", round(bond_npv, 2), "Discounted at 10%"],
    ["", "", ""],
    ["COMPARISON", "", ""],
    ["Nominal Saving (Bond vs Loan)", nom_save, ""],
    ["NPV Saving (Bond vs Loan)", round(npv_save, 2), ""],
    ["Saving %", round(save_pct, 4), f"~{save_pct*100:.2f}%"],
    ["Recommendation", "Structured Bond", ""],
]
for r in rows:
    ws.append(r)

# Cashflow schedule sheet
cf = wb.create_sheet("CashflowSchedule")
cf.append(["Year", "Loan Cashflow", "Bond Cashflow", "Loan DF", "Bond DF"])
cf.append([0, loan_fee, BOND_ISS, loan_fee, BOND_ISS])
for t in range(1, T + 1):
    df = (1 + DR) ** t
    cf.append([t, loan_annual, bond_annual, round(loan_annual / df, 2), round(bond_annual / df, 2)])
cf.append(["Total", loan_total, bond_total, round(loan_npv, 2), round(bond_npv, 2)])

# Formatting
title_font = Font(bold=True, size=14, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F4E78")
section_font = Font(bold=True, size=11, color="FFFFFF")
section_fill = PatternFill("solid", fgColor="2E75B6")
bold = Font(bold=True)

ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws.merge_cells("A1:C1")
ws["A2"].font = Font(italic=True, size=11)
ws.merge_cells("A2:C2")

for cell_ref in ["A4", "A15", "A25", "A35"]:
    ws[cell_ref].font = section_font
    ws[cell_ref].fill = section_fill

for row in [4, 15, 25, 35]:
    for col in ["B", "C"]:
        ws[f"{col}{row}"].font = section_font
        ws[f"{col}{row}"].fill = section_fill

# Number formatting for value column
zar_rows = list(range(5, 14)) + list(range(16, 24)) + list(range(26, 34)) + [36, 37]
for r in zar_rows:
    c = ws.cell(row=r, column=2)
    if isinstance(c.value, (int, float)):
        if c.value < 1 and c.value > 0:
            c.number_format = "0.00%"
        else:
            c.number_format = '"R" #,##0'

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 32

# Bold totals
for r in [22, 23, 32, 33, 36, 37, 38, 39]:
    ws.cell(row=r, column=1).font = bold
    ws.cell(row=r, column=2).font = bold

# Cashflow sheet formatting
cf["A1"].font = bold
for col in ["B", "C", "D", "E"]:
    cf.column_dimensions[col].width = 18
    cf[f"{col}1"].font = bold
for r in range(2, T + 4):
    for col in ["B", "C", "D", "E"]:
        cf[f"{col}{r}"].number_format = '"R" #,##0'

out = r"C:\Users\rudol\Documents\veilguard\CherryRain_CostEfficiency_Model.xlsx"
wb.save(out)
print("SAVED:", out)
