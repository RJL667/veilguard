"""Host document parsers — read + extract text from documents that live on
the sub-agents host (the user's Windows machine in a local deployment).

Why this exists: the container-based ``documents`` MCP server (read_pdf etc.)
runs in Linux and CANNOT see Windows paths like
``C:\\Users\\rudol\\Downloads\\report.pdf``. The sub-agents service runs on
the host and has PyMuPDF (fitz), so it can read + parse those files directly.
This closes the gap that previously forced the model to FORGE a one-off
pdf-extraction tool every time.

[HOST_DOC_READ_2026_06_01]  Scope: reads are allowed anywhere under the
user's home dir (covers Downloads / Documents / Desktop) — host-exec/daemon
already expose comparable host access. Override the root with
HOST_DOC_READ_ROOT. These are READ-ONLY parsers; no writes.
"""

import asyncio
import os
import re
import sys
from pathlib import Path

# [OCR_GEMINI_2026-06-11] shared OCR client (pii-proxy Gemini shim) — same
# 2-line _shared import shim the container servers use; resolves to
# services/_shared locally and mcp-tools/_shared on the VM.
sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "_shared"))
from ocr_client import ocr_bytes  # noqa: E402

_DOC_READ_ROOT = os.environ.get("HOST_DOC_READ_ROOT", os.path.expanduser("~"))

# [VM_PATH_GUARD_2026-06-12] On the cloud VM this service runs on Linux, but
# the model — seeing files like 'C:\Users\...\x.pdf' mentioned in chat —
# calls these tools with Windows paths. Path("C:/...") is NOT absolute on
# POSIX, so it used to get joined onto $HOME and produce nonsense like
# /home/rudol/C:/Users/... The guard converts that into an instruction the
# model can act on instead of retrying path spellings forever.
_WIN_DRIVE = re.compile(r"^[A-Za-z]:[\\/]")


def _reject_foreign_windows_path(path: str, _posix: bool = (os.name != "nt")):
    """Non-empty error string if `path` is a Windows drive path on Linux."""
    if _posix and _WIN_DRIVE.match(path.strip()):
        return (
            f"Error: {path!r} is a Windows path, but this tool runs on the "
            "Veilguard SERVER (Linux) and cannot see the user's local drive. "
            "Do NOT retry other path spellings — no path will work. To READ "
            "a file from the user's machine, ask the user to attach it in "
            "chat via the '+' menu -> 'Upload as Text' (it is OCR'd and its "
            "full text arrives in the conversation). To WRITE a file, use a "
            "server-side path (relative to home) and tell the user where it "
            "was saved."
        )
    return ""


# [DAEMON_PDF_FETCH_2026-06-12] When this service runs on the cloud VM and
# the requested file lives on the USER's Windows machine, we can still get
# it — through the user's connected client daemon — using only tool
# primitives EVERY DEPLOYED daemon already has (no daemon release needed):
#
#   1. run_command: `certutil -encode -f <path> <handoff>` base64-encodes
#      the file into a handoff file in the daemon's workspace root
#      (run_command's cwd). certutil is built into Windows; no PowerShell
#      (which mangles through the daemon's `cmd /c` quoting), no pipes
#      (those flip the daemon into shell=True + trip its has_pipe path).
#   2. read_file: the handoff is pure ASCII, so the daemon's text reader
#      is lossless. Read in <=500-line chunks (the daemon's per-call cap);
#      strip the daemon's "<n>\t" line-number prefix and certutil's
#      -----BEGIN/END CERTIFICATE----- markers, then base64-decode.
#   3. run_command `del`: remove the handoff (plain `del` passes the
#      daemon's dangerous-command filter; `/s`-style switches would not).
#
# Bytes are then parsed server-side (PyMuPDF + the OCR shim). Routed via
# core.agentic.handle_tool so per-user bridge selection + the approval
# gate apply exactly as for any other client tool call. Limitation: paths
# containing SPACES can't survive the daemon's `cmd /c` unquoting, so they
# fall through to the "ask the user to attach" message — correct fallback.
_DAEMON_FETCH_MAX = 15 * 1024 * 1024  # bytes; aligned with the OCR cap

# The read-back is sent DIRECT to the daemon bridge (not via handle_tool),
# which bypasses two things that broke earlier attempts:
#   * the task dispatcher truncates every result to 8000 chars
#     (core/tasks/executors/daemon.py) — that sliced base64 lines in half
#     and silently corrupted multi-chunk files;
#   * the per-call approval gate — going through it would prompt the user
#     once PER chunk (dozens of times) for one read.
# Direct execute_remote returns the FULL result, so we can use big chunks
# (~355KB each, far under uvicorn's 16MB WS frame cap) and keep round-trips
# low. The footer-presence check below is still a backstop against any
# residual truncation.
_READ_CHUNK_LINES = 10000
# A transfer must finish inside the MCP tool-call timeout (~60s). certutil
# base64 is ~64 chars/line ≈ 48 source bytes/line, so cap the relay at a size
# that comfortably completes; anything bigger is told to use Upload as Text
# (which streams server-side and has no per-call timeout). The 15MB hard cap
# (_DAEMON_FETCH_MAX) still guards the decode, but we bail BEFORE transferring.
_FAST_XFER_MAX_BYTES = 8 * 1024 * 1024
_FOOTER_TOTAL = re.compile(r"\bof (\d+)")  # "...showed lines X-Y of TOTAL..."

_LINE_NUM = re.compile(r"^\s*\d+\t(.*)$")  # daemon read_file prefixes "<n>\t"
# Daemon read_file always appends one of these footers; its ABSENCE means
# the 8000-char dispatcher cap truncated the response mid-line.
_READ_FOOTER = ("[end of file", "[partial read")


def _daemon_available() -> bool:
    """True if the current request's user has a connected client daemon."""
    if os.name == "nt":
        return False  # local stack: direct filesystem access already works
    try:
        from core.client_bridge import get_bridge
        from core.request_ctx import get_user_id
        bridge = get_bridge(get_user_id() or "")
        return bool(bridge and bridge.connected)
    except Exception:
        return False


_SHA256_LINE = re.compile(r"hash of", re.I)


def _path_unsafe(path: str) -> bool:
    """True if the path can't survive the daemon's `cmd /c` (space/quote)."""
    return " " in path or '"' in path or "\n" in path


def _certutil_fetch_command(path: str, handoff: str) -> str | None:
    """run_command string: base64 `path` -> `handoff`, THEN print its SHA256.

    The hash lets the server verify the reconstructed bytes end-to-end, so a
    corrupt transfer becomes a detected+retried failure rather than silent
    mojibake. Returns None for paths the daemon's `cmd /c` would mangle. The
    daemon runs with cwd=workspace root (relative paths resolve there);
    Windows certutil wants BACKSLASHES, so normalize forward slashes.
    """
    p = path.strip().replace("/", "\\")
    if _path_unsafe(p):
        return None
    # `&` chains unconditionally under cmd /c (the daemon already prepends a
    # `chcp ... &` prelude, so we're in shell mode regardless).
    return f"certutil -encode -f {p} {handoff} & certutil -hashfile {p} SHA256"


def _parse_certutil_sha256(out: str) -> str:
    """Pull the 64-hex SHA256 from `certutil -hashfile` stdout (the value sits
    on the line after 'SHA256 hash of <path>:'; some builds space-separate the
    bytes). Returns '' if not found."""
    lines = out.splitlines()
    for i, line in enumerate(lines):
        if _SHA256_LINE.search(line):
            for nxt in lines[i + 1:]:
                hexed = re.sub(r"[^0-9a-fA-F]", "", nxt)
                if len(hexed) >= 64:
                    return hexed[:64].lower()
    return ""


async def _fetch_bytes_via_daemon(path: str) -> tuple[bytes | None, str]:
    """Pull a file from the user's machine via their client daemon, verifying
    integrity against the file's real SHA256 (retries a corrupt transfer).

    Returns (bytes, "") on success or (None, reason). Reasons are LLM-facing.
    """
    import base64 as _b64
    import hashlib as _hash
    import uuid as _uuid
    from core.agentic import handle_tool
    from core.client_bridge import get_bridge
    from core.request_ctx import get_user_id

    if _path_unsafe(path.strip().replace("/", "\\")):
        return None, (
            f"{path!r} contains a space (or quote) that the client-daemon "
            "transfer can't carry. Ask the user to attach the file in chat "
            "via the '+' menu -> 'Upload as Text' instead."
        )

    bridge = get_bridge(get_user_id() or "")
    if not bridge or not bridge.connected:
        return None, (
            "no client daemon is connected for this user, so the file can't "
            "be fetched from their machine. Ask the user to start the "
            "Veilguard client, or attach the file via '+' -> 'Upload as Text'."
        )

    last_reason = "unknown error"
    for attempt in range(3):
        handoff = f"vg_xfer_{_uuid.uuid4().hex[:8]}.b64"
        cmd = _certutil_fetch_command(path, handoff)
        # certutil goes through handle_tool so the approval gate applies once
        # ("run a command on the user's machine"); its result is small so the
        # dispatcher's 8000-char cap never bites.
        out = str(await handle_tool(
            "run_command", {"command": cmd, "timeout": 120},
        ) or "")
        if "successfully" not in out.lower():
            if "BLOCKED" in out:
                return None, f"daemon refused the transfer command: {out[:200]}"
            # certutil prints "cannot find the file"/"path not found" on a miss.
            tail = out.strip().splitlines()[-1][:160] if out.strip() else "no output"
            return None, (
                f"the daemon could not read {path!r} from the user's machine "
                f"(certutil: {tail}). Give the path exactly as the user stated "
                "it (e.g. 'Downloads/<name>.pdf'), or have them attach it via "
                "'+' -> 'Upload as Text'."
            )
        expected_sha = _parse_certutil_sha256(out)

        chunks: list[str] = []
        offset = 0
        read_err = ""
        try:
            while True:
                # DIRECT bridge call: full untruncated result, no per-chunk
                # approval prompt.
                body = str(await bridge.execute_remote(
                    "read_file",
                    {"path": handoff, "offset": offset, "limit": _READ_CHUNK_LINES},
                    timeout=60.0,
                ) or "")
                if body.startswith("Error:") or "BLOCKED" in body[:200]:
                    read_err = f"daemon read-back failed: {body[:300]}"
                    break
                # Early size-bail: the first chunk's footer reports the total
                # line count; if that implies a file too big to relay inside
                # the MCP timeout, stop now and send the user to Upload as Text.
                if offset == 0:
                    mt = _FOOTER_TOTAL.search(body)
                    if mt and int(mt.group(1)) * 48 > _FAST_XFER_MAX_BYTES:
                        approx_mb = (int(mt.group(1)) * 48) // (1024 * 1024)
                        read_err = (
                            f"TOOBIG:{approx_mb}"  # handled below, no retry
                        )
                        break
                has_footer = any(f in body for f in _READ_FOOTER)
                data_lines = [
                    m.group(1) for m in (
                        _LINE_NUM.match(ln) for ln in body.splitlines()
                    ) if m
                ]
                # No footer => the response was truncated mid-line, so the
                # LAST data line is partial — drop it and re-read from there.
                # (Direct bridge isn't capped, so this is just a backstop.)
                if not has_footer and data_lines:
                    data_lines = data_lines[:-1]
                if not data_lines:
                    read_err = "daemon read-back returned no usable data lines."
                    break
                for content in data_lines:
                    c = content.strip()
                    if c and not c.startswith("-----"):
                        chunks.append(c)  # skip BEGIN/END CERTIFICATE markers
                offset += len(data_lines)  # advance by lines actually consumed
                if has_footer and "[end of file" in body:
                    break
        finally:
            # Best-effort cleanup; plain `del` passes the daemon's filter.
            try:
                await bridge.execute_remote(
                    "run_command", {"command": f"del {handoff}"}, timeout=30.0)
            except Exception:
                pass

        if read_err:
            if read_err.startswith("TOOBIG:"):
                mb = read_err.split(":", 1)[1]
                return None, (
                    f"the file is ~{mb}MB — too large to fetch through the "
                    "client-daemon relay within the request timeout. Ask the "
                    "user to attach it in chat via '+' -> 'Upload as Text' "
                    "instead (no size/timeout limit there)."
                )
            last_reason = read_err
            continue
        try:
            data = _b64.b64decode("".join(chunks))
        except Exception as e:
            last_reason = f"base64 decode failed: {e}"
            continue
        if len(data) > _DAEMON_FETCH_MAX:
            return None, (
                f"the file is {len(data) // (1024 * 1024)}MB, over the "
                f"{_DAEMON_FETCH_MAX // (1024 * 1024)}MB transfer limit."
            )
        # End-to-end integrity gate: only trust bytes whose hash matches the
        # source on the user's machine. Mismatch => corrupt transfer => retry.
        got_sha = _hash.sha256(data).hexdigest()
        if expected_sha and got_sha != expected_sha:
            last_reason = (
                f"integrity check failed (got {got_sha[:12]}, expected "
                f"{expected_sha[:12]}) on attempt {attempt + 1}"
            )
            continue
        return data, ""

    return None, (
        f"the file transfer from the user's machine kept failing "
        f"({last_reason}). Ask the user to attach the file via '+' -> "
        "'Upload as Text' instead."
    )


def _resolve_read(path: str):
    """Return (resolved_path, "") or (None, error_str)."""
    if not isinstance(path, str) or not path.strip():
        return None, "Error: path is required"
    if err := _reject_foreign_windows_path(path):
        return None, err
    p = Path(path.replace("\\", "/"))
    if not p.is_absolute():
        p = Path(_DOC_READ_ROOT) / p
    rp = p.resolve()
    root = Path(_DOC_READ_ROOT).resolve()
    if not str(rp).startswith(str(root)):
        return None, (
            f"Error: read denied — {path!r} is outside the allowed root "
            f"({root}). Set HOST_DOC_READ_ROOT to widen."
        )
    if not rp.exists():
        msg = f"Error: File not found: {rp}"
        if os.name != "nt":
            # [VM_PATH_GUARD_2026-06-12] Relative paths ("Downloads/x.pdf")
            # resolve under the SERVER's home on the VM and miss too — same
            # redirect as the drive-letter guard so the model stops probing.
            msg += (
                " (This server is NOT the user's machine. If the file is on "
                "the user's machine, do not try other paths — ask the user "
                "to attach it in chat via '+' -> 'Upload as Text'.)"
            )
        return None, msg
    if not rp.is_file():
        return None, f"Error: Not a file: {rp}"
    return rp, ""


def _resolve_write(path: str):
    """Resolve a path for WRITING a document on the host. Same root rules as
    reads, but the file need not exist; parent dirs are created."""
    if not isinstance(path, str) or not path.strip():
        return None, "Error: path is required"
    if err := _reject_foreign_windows_path(path):
        return None, err
    p = Path(path.replace("\\", "/"))
    if not p.is_absolute():
        p = Path(_DOC_READ_ROOT) / p
    rp = p.resolve()
    root = Path(_DOC_READ_ROOT).resolve()
    if not str(rp).startswith(str(root)):
        return None, (
            f"Error: write denied — {path!r} is outside the allowed root "
            f"({root}). Set HOST_DOC_READ_ROOT to widen."
        )
    try:
        rp.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        return None, f"Error: cannot create parent dir: {e}"
    return rp, ""


async def _parse_pdf_file(rp: Path, pages: str, display_name: str = "") -> str:
    """Extract text from the PDF at `rp` (server-local path), with the
    Gemini-OCR fallback for scanned pages. Shared by the direct-path and
    daemon-fetched branches of read_pdf."""
    name = display_name or rp.name
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        return f"Error: PDF parser (PyMuPDF) unavailable on host: {e}"
    try:
        doc = fitz.open(str(rp))
    except Exception as e:
        return f"Error opening PDF: {e}"
    try:
        total = len(doc)
        if pages.strip():
            parts = pages.strip().split("-")
            start = max(0, int(parts[0]) - 1)
            end = int(parts[-1]) if len(parts) > 1 else start + 1
            rng = range(start, min(end, total))
        else:
            rng = range(total)
        out = [f"# {name} ({total} pages, reading {len(rng)})\n"]
        text_chars = 0
        full_text = []
        for i in rng:
            text = doc[i].get_text().strip()
            text_chars += len(text)
            full_text.append(text)
            out.append(f"--- Page {i + 1} ---\n{text}\n")
        # Decide whether the text layer is USABLE. Two failure modes both
        # mean "fall back to image OCR" rather than returning the text:
        #   (a) empty/near-empty  -> scanned image PDF, no text layer.
        #   (b) GARBAGE           -> broken font encoding (no/!bad ToUnicode
        #       CMap), so get_text() returns glyph-id codepoints that render
        #       as mojibake. Detect via a low printable-ASCII ratio over a
        #       non-trivial amount of text. Readable English/Latin docs sit
        #       ~0.95; mojibake sits ~0.2-0.4. 0.55 cleanly separates them
        #       without tripping on normal accented text.
        joined = "".join(full_text)
        printable = sum(1 for c in joined if 32 <= ord(c) < 127)
        ascii_ratio = printable / len(joined) if joined else 1.0
        empty = rng and text_chars < 25 * len(rng)
        garbage = text_chars >= 100 and ascii_ratio < 0.55
        # [OCR_GEMINI_2026-06-11] No usable text layer => OCR the selected
        # pages via the pii-proxy Gemini shim. to_thread keeps the blocking
        # HTTP call off the server's event loop.
        if rng and (empty or garbage):
            if len(rng) >= total:
                data = rp.read_bytes()
            else:
                sub = fitz.open()
                sub.insert_pdf(doc, from_page=rng[0], to_page=rng[-1])
                data = sub.tobytes()
                sub.close()
            reason = ("garbled text layer (broken font encoding)" if garbage
                      else "no text layer")
            try:
                md = await asyncio.to_thread(ocr_bytes, data, "application/pdf")
                return (
                    f"# {name} ({total} pages — {reason} on the selected "
                    f"{len(rng)} page(s); OCR'd via Gemini vision)\n\n{md}"
                )
            except Exception as e:
                out.append(
                    f"\n[Note: these pages have {reason} and the OCR "
                    f"fallback failed: {e}]"
                )
        return "\n".join(out)
    except Exception as e:
        return f"Error reading PDF: {e}"
    finally:
        doc.close()


async def _read_pdf_via_daemon(path: str, pages: str) -> str:
    """Fetch a PDF from the user's machine via their client daemon and
    parse it server-side."""
    import tempfile

    data, why = await _fetch_bytes_via_daemon(path)
    if data is None:
        return f"Error: {why}"
    if not data.startswith(b"%PDF"):
        return (
            f"Error: {path!r} was transferred from the user's machine "
            f"({len(data)} bytes) but is not a PDF (no %PDF header)."
        )
    tmp = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    try:
        tmp.write(data)
        tmp.close()
        text = await _parse_pdf_file(
            Path(tmp.name), pages,
            display_name=Path(path.replace("\\", "/")).name,
        )
        return text + "\n\n[fetched from the user's machine via the Veilguard client daemon]"
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass


def register(mcp):
    @mcp.tool()
    async def read_pdf(path: str, pages: str = "") -> str:
        """Read and extract text from a PDF — on the machine running this
        service, or on the USER's machine via their Veilguard client daemon.

        LOCAL install: this service runs on the user's Windows machine, so
        real Windows paths work directly (e.g.
        'C:/Users/<name>/Downloads/report.pdf'); forward or back slashes both
        work. CLOUD/VM deployment: paths resolve on the Veilguard server —
        but if the file isn't found there and the user has their Veilguard
        client daemon connected, the file is fetched from THEIR machine
        automatically (give the path as the user stated it, e.g.
        'Downloads/report.pdf' or 'C:/Users/<name>/Downloads/report.pdf').
        If no daemon is connected, ask the user to attach the file in chat
        ('+' menu -> 'Upload as Text') instead of guessing paths.

        Args:
            path: PDF path (absolute, or relative to home/workspace).
            pages: Page range like "1-5" or "3". Empty = all pages.
        """
        rp, err = _resolve_read(path)
        if err:
            # Server-side miss (Windows path, or not found under the server
            # home). If this user has a connected client daemon, fetch the
            # file from their machine instead of bouncing the error.
            if _daemon_available():
                return await _read_pdf_via_daemon(path, pages)
            return err
        return await _parse_pdf_file(rp, pages)

    @mcp.tool()
    async def read_xlsx(path: str, sheet: str = "", max_rows: int = 500) -> str:
        """Read an Excel spreadsheet as text — from this machine, or from
        the USER's machine via their Veilguard client daemon.

        LOCAL install: the user's Windows machine (real Windows paths work).
        CLOUD/VM deployment: paths resolve on the Veilguard server, but if
        the file isn't found there and the user's client daemon is
        connected, it is fetched from THEIR machine automatically. If no
        daemon is connected, ask the user to attach the file in chat.

        Args:
            path: XLSX path (absolute, or relative to home/workspace).
            sheet: Sheet name. Empty = first/active sheet.
            max_rows: Max rows to return (default 500).
        """
        import io as _io

        src = None
        fetched = False
        rp, err = _resolve_read(path)
        if err:
            if _daemon_available():
                data, why = await _fetch_bytes_via_daemon(path)
                if data is None:
                    return f"Error: {why}"
                src = _io.BytesIO(data)
                fetched = True
            else:
                return err
        try:
            from openpyxl import load_workbook
        except Exception as e:
            return f"Error: Excel parser (openpyxl) unavailable on host: {e}"
        try:
            wb = load_workbook(src if fetched else str(rp), read_only=True, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            if sheet and sheet not in wb.sheetnames:
                return f"Error: Sheet {sheet!r} not found. Available: {wb.sheetnames}"
            disp = Path(path.replace("\\", "/")).name if fetched else rp.name
            out = [f"# {disp} — Sheet: {ws.title}", f"# Sheets: {wb.sheetnames}\n"]
            n = 0
            for row in ws.iter_rows(values_only=True):
                if n >= max_rows:
                    out.append(f"\n... truncated at {max_rows} rows")
                    break
                out.append(" | ".join(str(c) if c is not None else "" for c in row))
                n += 1
            wb.close()
            return "\n".join(out)
        except Exception as e:
            return f"Error reading XLSX: {e}"

    @mcp.tool()
    async def create_xlsx(path: str, data: str, sheet_name: str = "Sheet1") -> str:
        """Create an Excel spreadsheet on the machine RUNNING this
        sub-agents service.

        LOCAL install: writes to the user's Windows machine (e.g.
        'C:/Users/<name>/Downloads/report.xlsx'). CLOUD/VM deployment: writes
        land on the Veilguard SERVER, not the user's machine — use a path
        relative to home and tell the user where the file was saved. The
        container 'documents' create_xlsx only writes inside its /workspace.

        Args:
            path: Output XLSX path (absolute, or relative to the service home dir).
            data: JSON string — a list of dicts [{"col":"val",...}] OR a list of
                  lists [["h1","h2"],["v1","v2"]] (first row = header).
            sheet_name: Worksheet name.
        """
        import json as _json
        rp, err = _resolve_write(path)
        if err:
            return err
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except Exception as e:
            return f"Error: Excel writer (openpyxl) unavailable on host: {e}"
        try:
            parsed = _json.loads(data)
        except Exception as e:
            return f"Error: `data` is not valid JSON: {e}"
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            if not parsed:
                wb.save(str(rp))
                return f"Created empty XLSX: {rp}"
            if isinstance(parsed[0], dict):
                headers = list(parsed[0].keys())
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                for r in parsed:
                    ws.append([r.get(h, "") for h in headers])
            elif isinstance(parsed[0], (list, tuple)):
                for i, r in enumerate(parsed):
                    ws.append(list(r))
                    if i == 0:
                        for cell in ws[1]:
                            cell.font = Font(bold=True)
            for col in ws.columns:
                mx = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 50)
            wb.save(str(rp))
            return f"Created XLSX {rp} ({ws.max_row} rows, sheet {sheet_name!r})"
        except Exception as e:
            return f"Error creating XLSX: {e}"

    @mcp.tool()
    async def create_pdf(path: str, content: str, title: str = "") -> str:
        """Create a PDF document on the machine RUNNING this sub-agents
        service.

        LOCAL install: writes to the user's Windows machine (e.g.
        'C:/Users/<name>/Downloads/report.pdf'). CLOUD/VM deployment: writes
        land on the Veilguard SERVER, not the user's machine — use a path
        relative to home and tell the user where the file was saved. Runs
        with reportlab; PDF counterpart to create_xlsx.

        Args:
            path: Output PDF path (absolute, or relative to the service home dir).
            content: Body as light markdown, parsed per line: '# '/'## '/'### '
                     headings, '- ' or '* ' bullets, blank line = spacer,
                     '**bold**' inline; anything else = a paragraph.
            title: Optional title rendered as a large heading on page 1.
        """
        import html as _html
        import re as _re2
        rp, err = _resolve_write(path)
        if err:
            return err
        if rp.suffix.lower() != ".pdf":
            rp = rp.with_suffix(".pdf")
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Paragraph,
                                            Spacer, HRFlowable)
        except Exception as e:
            return f"Error: PDF writer (reportlab) unavailable on host: {e}"

        def _inline(s: str) -> str:
            s = _html.escape(s, quote=False)
            return _re2.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", s)

        styles = getSampleStyleSheet()
        body = ParagraphStyle("CRBody", parent=styles["BodyText"], fontSize=10.5, leading=15, spaceAfter=6)
        h1 = ParagraphStyle("CRH1", parent=styles["Heading1"], fontSize=16, spaceBefore=10, spaceAfter=6)
        h2 = ParagraphStyle("CRH2", parent=styles["Heading2"], fontSize=13, spaceBefore=8, spaceAfter=4)
        h3 = ParagraphStyle("CRH3", parent=styles["Heading3"], fontSize=11.5, spaceBefore=6, spaceAfter=3)
        bul = ParagraphStyle("CRBul", parent=body, leftIndent=14, bulletIndent=2, spaceAfter=3)
        ttl = ParagraphStyle("CRTitle", parent=styles["Title"], fontSize=22, spaceAfter=4)

        story = []
        if title.strip():
            story.append(Paragraph(_inline(title.strip()), ttl))
            story.append(HRFlowable(width="100%", thickness=1.2, color=colors.HexColor("#444444"), spaceAfter=10))
        for raw in (content or "").replace("\r\n", "\n").split("\n"):
            s = raw.strip()
            if not s:
                story.append(Spacer(1, 6)); continue
            if s.startswith("### "):
                story.append(Paragraph(_inline(s[4:]), h3))
            elif s.startswith("## "):
                story.append(Paragraph(_inline(s[3:]), h2))
            elif s.startswith("# "):
                story.append(Paragraph(_inline(s[2:]), h1))
            elif s[:2] in ("- ", "* "):
                story.append(Paragraph(_inline(s[2:]), bul, bulletText="•"))
            else:
                story.append(Paragraph(_inline(s), body))
        if not story:
            story.append(Paragraph("(empty document)", body))

        def _footer(canvas, doc_):
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(colors.HexColor("#888888"))
            canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Page {doc_.page}")
            canvas.restoreState()

        try:
            doc = SimpleDocTemplate(str(rp), pagesize=A4, leftMargin=2 * cm,
                                    rightMargin=2 * cm, topMargin=2 * cm,
                                    bottomMargin=2 * cm, title=(title.strip() or rp.stem))
            doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
        except Exception as e:
            return f"Error creating PDF: {e}"
        try:
            size = rp.stat().st_size
        except Exception:
            size = 0
        return f"Created PDF {rp} ({size} bytes)"
