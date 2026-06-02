
import openpyxl
from openpyxl.utils import get_column_letter

def run(path: str, action: str) -> str:
    try:
        if action == "create":
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            
            # Add headers
            headers = ["Name", "Score", "Status"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add sample data
            data = [
                ["Alice", 95, "Pass"],
                ["Bob", 87, "Pass"],
                ["Charlie", 72, "Pass"],
                ["Diana", 58, "Fail"]
            ]
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(path)
            return f"Excel file created at {path}"
        
        elif action == "read":
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            
            result = f"Sheet: {ws.title}\n"
            result += "=" * 40 + "\n"
            
            for row in ws.iter_rows(values_only=True):
                result += " | ".join(str(v) if v is not None else "" for v in row) + "\n"
            
            return result
        
        else:
            return f"Unknown action: {action}"
    
    except Exception as e:
        return f"Error: {str(e)}"
