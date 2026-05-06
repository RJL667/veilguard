import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Model"

bold = Font(bold=True)
hdr = Font(bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="1F4E78")
sec = PatternFill("solid", fgColor="D9E1F2")

ws["A1"] = "Cherry Rain — Cost-Efficiency Model"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:C1")

rows = [
    ("", "", ""),
    ("Inputs", "", "Section"),
    ("Endpoints", 250, ""),
    ("Avg phishing incidents / month (pre)", 18, ""),
    ("Avg cost per incident (ZAR)", 42000, ""),
    ("Phishield monthly fee (ZAR)", 38500, ""),
    ("Implementation one-off (ZAR)", 65000, ""),
    ("Expected reduction in incidents", 0.78, ""),
    ("", "", ""),
    ("Monthly Calculations", "", "Section"),
    ("Pre-Phishield monthly loss (ZAR)", "=B4*B5", ""),
    ("Post-Phishield incidents", "=B4*(1-B8)", ""),
    ("Post-Phishield monthly loss (ZAR)", "=B12*B5", ""),
    ("Gross monthly saving (ZAR)", "=B11-B13", ""),
    ("Net monthly saving (ZAR)", "=B14-B6", ""),
    ("", "", ""),
    ("Annual View", "", "Section"),
    ("Annual gross saving (ZAR)", "=B14*12", ""),
    ("Annual Phishield fee (ZAR)", "=B6*12", ""),
    ("Annual net saving (ZAR)", "=B15*12-B7", ""),
    ("ROI year 1", "=B20/(B19+B7)", ""),
    ("Payback (months)", "=B7/B15", ""),
]

for i, (a, b, tag) in enumerate(rows, start=2):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
    if tag == "Section":
        ws.cell(row=i, column=1).font = bold
        ws.cell(row=i, column=1).fill = sec
        ws.cell(row=i, column=2).fill = sec

zar = '"R" #,##0'
for r in (5, 6, 7, 8, 12, 14, 15, 16, 19, 20, 21):
    ws.cell(row=r, column=2).number_format = zar
ws["B9"].number_format = "0.0%"
ws["B22"].number_format = "0.0%"
ws["B23"].number_format = "0.00"

ws.column_dimensions["A"].width = 42
ws.column_dimensions["B"].width = 18

wb.save(r"C:\Users\rudol\Documents\veilguard\CherryRain_CostEfficiency_Model.xlsx")
print("OK")
